"""
Build nztcs.json from NZTCS Exported Data.xlsx

Produces a flat lookup keyed by lowercase scientific name.
Each entry uses the most-recently-assessed record (by Year Assessed).
Aliases (Previous Name, Alternative Names) also get index entries.

Output: nztcs.json (same directory as this script)
"""

import sys, json, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

XLSX = r'C:\Users\User\Downloads\NZTCS Exported Data.xlsx'
OUT  = r'C:\Users\User\GBIF-Record-Finder\.claude\worktrees\infallible-babbage-f27395\nztcs.json'

def clean(s):
    """Return stripped string or None."""
    if s is None:
        return None
    s = str(s).strip()
    return s if s else None

def norm_name(s):
    """
    Normalise a scientific name for lookup:
    - lowercase
    - strip leading/trailing quotes and whitespace
    - collapse internal whitespace
    """
    if not s:
        return None
    s = s.strip().strip('"').strip("'").lower()
    s = re.sub(r'\s+', ' ', s)
    # Strip authority — keep only genus + specific epithet (first 2 tokens)
    # BUT only if it looks like a binomial (2+ words). Uninomials stay as-is.
    tokens = s.split()
    if len(tokens) >= 2:
        # Heuristic: authority starts at first token that's not a lowercase word
        # Actually let's keep up to 3 tokens (subsp.) — GBIF names are usually 2
        # For robustness, just keep all tokens — we want to match exactly what GBIF returns
        pass
    return s

def split_alts(s):
    """Split Alternative Names field (comma or semicolon separated)."""
    if not s:
        return []
    parts = re.split(r'[;,]', str(s))
    return [p.strip() for p in parts if p.strip()]

print("Loading workbook…")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb.active

headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print(f"  Columns: {len(headers)}")

# Main lookup: key → entry dict
# We keep the most recent assessment per species name
lookup = {}   # lowercase name → entry
aliases = {}  # lowercase alias → canonical lowercase name

rows_read = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    rows_read += 1
    d = dict(zip(headers, row))

    current_name = clean(d.get('Current Species Name'))
    if not current_name:
        continue

    year = d.get('Year Assessed')
    try:
        year = int(year) if year else None
    except (ValueError, TypeError):
        year = None

    entry = {
        'name':     current_name,
        'common':   clean(d.get('Preferred Common Name')),
        'maori':    clean(d.get('Preferred Māori Name')),
        'status':   clean(d.get('Status')),
        'category': clean(d.get('Category')),
        'year':     year,
        'bioStatus': clean(d.get('Bio Status')),
        'order':    clean(d.get('Order')),
        'family':   clean(d.get('Family')),
    }

    key = norm_name(current_name)
    if key:
        # Keep newer assessment if duplicate
        existing = lookup.get(key)
        if not existing or (year and (not existing['year'] or year > existing['year'])):
            lookup[key] = entry

        # Previous name → alias
        prev = clean(d.get('Previous Name'))
        if prev:
            pk = norm_name(prev)
            if pk and pk != key:
                aliases[pk] = key

        # Alternative names → aliases
        for alt in split_alts(d.get('Alternative Names')):
            ak = norm_name(alt)
            if ak and ak != key:
                aliases[ak] = key

wb.close()
print(f"  Rows read: {rows_read}")
print(f"  Unique names: {len(lookup)}")
print(f"  Aliases: {len(aliases)}")

# Merge aliases that don't shadow a real name
added = 0
for alias_key, canonical_key in aliases.items():
    if alias_key not in lookup and canonical_key in lookup:
        lookup[alias_key] = lookup[canonical_key]
        added += 1

print(f"  Alias entries added: {added}")

# ── Subspecies binomial aliases ────────────────────────────────────────────
# GBIF often records observations at species level (e.g. "Egretta sacra")
# even when NZTCS only lists the subspecies (e.g. "Egretta sacra sacra").
# Build a binomial alias (first 2 words) for every 3-word Latin subspecies
# entry that doesn't already have a binomial key.  When multiple subspecies
# exist for the same species, pick the most-threatened one.

STATUS_RANK = {
    'Nationally Critical': 0, 'Nationally Endangered': 1,
    'Nationally Vulnerable': 2, 'Declining': 3,
    'Naturally Uncommon': 4, 'Relict': 5, 'Uncommon': 6,
    'Recovering': 7, 'Nationally Increasing': 8, 'Extinct': 9,
    'Not Threatened': 10, 'Introduced and Naturalised': 11,
    'Vagrant': 12, 'Data Deficient': 13,
}

def is_latin_subsp(key, entry):
    """True if key looks like a 3-word Latin trinomial (genus epithet subsp)."""
    parts = key.split()
    if len(parts) != 3:
        return False
    # All three parts should be simple alphabetic tokens (allow hyphens)
    if not all(re.match(r'^[a-z][a-z\-]*$', p) for p in parts):
        return False
    # Original name's first letter should be uppercase (Latin genus)
    orig = entry.get('name', '')
    return orig[:1].isupper()

from collections import defaultdict
binomial_candidates = defaultdict(list)   # binomial → list of entries

for key, entry in lookup.items():
    if is_latin_subsp(key, entry):
        parts = key.split()
        binomial = parts[0] + ' ' + parts[1]
        binomial_candidates[binomial].append(entry)

subsp_added = 0
for binomial, entries in binomial_candidates.items():
    if binomial in lookup:
        continue  # binomial already has its own entry — don't overwrite
    # Pick the most-threatened entry for this binomial
    best = min(entries, key=lambda e: (STATUS_RANK.get(e.get('status'), 99), -(e.get('year') or 0)))
    lookup[binomial] = best
    subsp_added += 1

print(f"  Subspecies binomial aliases added: {subsp_added}")

# ── Old-name binomial aliases ──────────────────────────────────────────────
# Many NZ taxa were reclassified to a new genus (e.g. Porzana → Zapornia).
# The Previous Name field captures the old name, e.g. "Porzana tabuensis tabuensis
# Gmelin, 1789". The alias dict maps the normalised old name (including authority)
# to the canonical current key — but GBIF still records the old binomial
# "Porzana tabuensis" which is never in the lookup.
#
# Fix: for every alias key whose first two tokens form a binomial not yet in the
# lookup, add that binomial pointing to the same entry.

def looks_like_latin_binomial_prefix(key):
    """True if the key starts with two lowercase alphabetic words (Latin binomial)."""
    parts = key.split()
    if len(parts) < 2:
        return False
    return bool(re.match(r'^[a-z][a-z\-]+$', parts[0]) and re.match(r'^[a-z][a-z\-]+$', parts[1]))

old_binomial_added = 0
# aliases still maps alias_key → canonical_key (before merging into lookup)
# But we already merged them into lookup above. Re-derive from lookup keys that
# aren't in the original lookup (i.e. came from aliases) — simpler: iterate
# all existing lookup keys and generate binomials for any that look like old names.

# Collect all lookup keys that look like "genus species subsp authority..."
# (4+ tokens) and generate binomial aliases from them.
for key in list(lookup.keys()):
    parts = key.split()
    # Must have at least 3 tokens and start with two lowercase alpha words
    if len(parts) < 3:
        continue
    if not looks_like_latin_binomial_prefix(key):
        continue
    binomial = parts[0] + ' ' + parts[1]
    if binomial in lookup:
        continue
    # Add binomial alias pointing to this entry
    lookup[binomial] = lookup[key]
    old_binomial_added += 1

print(f"  Old-name binomial aliases added: {old_binomial_added}")
print(f"  Total lookup entries: {len(lookup)}")

# Status summary
from collections import Counter
status_counts = Counter(v['status'] for v in lookup.values())
print("\nStatus distribution (unique entries):")
for status, count in status_counts.most_common():
    print(f"  {status}: {count}")

# Write JSON
print(f"\nWriting {OUT}…")
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(lookup, f, ensure_ascii=False, separators=(',', ':'))

import os
size_kb = os.path.getsize(OUT) / 1024
print(f"Done. File size: {size_kb:.1f} KB")
